import os
import logging

from PyPDF2 import PdfReader
from openpyxl import load_workbook
from docx import Document

from app.config import abspath

def open_txt_file(file_path: str) -> str:
    """
    Opens a text file and returns its content as a string.

    Args:
        file_path: The path to the text file.
    Returns:
        The content of the text file as a string.
    """
    try:
        with open(file_path, "r", encoding="utf-8") as file:
            return file.read()
    except Exception as e:
        logging.error(f"Error al abrir el archivo de texto: {e}")
        return None


def create_txt_file(file_name: str, content: str) -> str:
    """
    Creates a text file with the given content.

    Args:
        file_name: The name of the file to be created.
        content: The content to be written to the file.
    Returns:
        The path to the created file.
    """
    try:
        destination_file_path = os.path.join(abspath, f"temp/{file_name}")

        with open(file=destination_file_path, mode="w", encoding="utf-8") as file:
            file.write(content)
            return destination_file_path
    except Exception as e:
        logging.error(f"Error al crear el archivo de texto: {e}")
        return None
    

def pdf_to_txt(pdf_path: str) -> str:
    """Converts a PDF file to a text string.
    
    Args:
        pdf_path: The path to the PDF file.
    Returns:
        A string containing the text from the PDF file.
    """
    text = ""
    try:
        reader = PdfReader(pdf_path)
        for i, page in enumerate(reader.pages):
            page_text = page.extract_text() or ""
            text += f"\n=== PAGE {i + 1} ===\n{page_text}\n"

        return text
    except Exception as e:
        logging.error(f"Error al leer el archivo PDF: {e}")
    return text


def xlxs_to_txt(xlxs_path: str) -> str:
    """Converts an Excel file to a text string.
    
    Args:
        xlxs_path: The path to the Excel file.
    Returns:
        A string containing the text from the Excel file.
    """
    text = ""
    try:
        workbook = load_workbook(xlxs_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                row_text += "\t".join([str(cell) if cell is not None else "" for cell in row])
                text += row_text + "\n"
        return text
    except Exception as e:
        logging.error(f"Error al leer el archivo Excel: {e}")
    return text


def docx_to_txt(doc_path: str) -> str:
    """Converts a Word document to a text string.
    
    Args:
        doc_path: The path to the Word document.
    Returns:
        A string containing the text from the Word document.
    """
    text = ""
    try:
        document = Document(doc_path)
        for paragraph in document.paragraphs:
            text += paragraph.text + "\n"
        return text
    except Exception as e:
        logging.error(f"Error al leer el archivo Word: {e}")
    return text